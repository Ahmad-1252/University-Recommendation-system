"""
Data Exporter - Save university programs to CSV, Excel, and MongoDB
Supports multiple export formats and database persistence
"""

import csv
import json
import logging
import os
from datetime import datetime
from typing import Any, Dict, List

from dotenv import load_dotenv
from pymongo import MongoClient

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("data_export.log"),
    ],
)
logger = logging.getLogger(__name__)


class UniversityDataExporter:
    """Export university program data to multiple formats"""

    def __init__(self):
        """Initialize exporter with database connection"""
        self.mongo_uri = os.getenv("MONGODB_URI", "mongodb://localhost:27017/")
        self.mongo_client = None
        self.db = None
        self.collection = None
        self.export_dir = "exported_data"

        # Create export directory
        if not os.path.exists(self.export_dir):
            os.makedirs(self.export_dir)
            logger.info(f"Created export directory: {self.export_dir}")

    def connect_mongodb(self) -> bool:
        """Connect to MongoDB database"""
        try:
            self.mongo_client = MongoClient(
                self.mongo_uri, serverSelectionTimeoutMS=5000
            )
            self.mongo_client.admin.command("ping")
            self.db = self.mongo_client["university_scraper"]
            self.collection = self.db["programs"]
            logger.info("[OK] MongoDB connection established")
            return True
        except Exception as e:
            logger.error(f"[ERROR] Failed to connect to MongoDB: {e}")
            return False

    def export_to_csv(
        self, programs: List[Dict[str, Any]], filename: str = None
    ) -> bool:
        """Export programs to CSV file"""
        try:
            if filename is None:
                filename = f"programs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

            filepath = os.path.join(self.export_dir, filename)

            if not programs:
                logger.warning("[WARNING] No programs to export to CSV")
                return False

            # Define CSV columns
            fieldnames = [
                "university_name",
                "program_name",
                "degree_type",
                "duration",
                "tuition_fees",
                "admission_requirements",
                "language_requirements",
                "application_deadline",
                "program_url",
                "country",
                "city",
                "ranking",
                "description",
                "confidence_score",
                "extracted_at",
            ]

            # Write CSV
            with open(filepath, "w", newline="", encoding="utf-8") as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()

                for program in programs:
                    # Filter only relevant fields
                    row = {field: program.get(field, "") for field in fieldnames}
                    # Convert datetime to string if needed
                    if row["extracted_at"]:
                        row["extracted_at"] = str(row["extracted_at"])
                    writer.writerow(row)

            logger.info(f"[OK] Exported {len(programs)} programs to CSV: {filepath}")
            return True

        except Exception as e:
            logger.error(f"[ERROR] Failed to export to CSV: {e}")
            return False

    def export_to_excel(
        self, programs: List[Dict[str, Any]], filename: str = None
    ) -> bool:
        """Export programs to Excel file with formatting"""
        try:
            if not OPENPYXL_AVAILABLE:
                logger.warning(
                    "[WARNING] openpyxl not installed. Install with: pip install openpyxl"
                )
                return False

            if filename is None:
                filename = f"programs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            filepath = os.path.join(self.export_dir, filename)

            if not programs:
                logger.warning("[WARNING] No programs to export to Excel")
                return False

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Programs"

            # Define columns
            columns = [
                ("University", "university_name"),
                ("Program Name", "program_name"),
                ("Degree Type", "degree_type"),
                ("Duration", "duration"),
                ("Tuition Fees", "tuition_fees"),
                ("Admission Requirements", "admission_requirements"),
                ("Language Requirements", "language_requirements"),
                ("Application Deadline", "application_deadline"),
                ("Program URL", "program_url"),
                ("Country", "country"),
                ("City", "city"),
                ("Ranking", "ranking"),
                ("Description", "description"),
                ("Confidence Score", "confidence_score"),
                ("Extracted At", "extracted_at"),
            ]

            # Write header with formatting
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for col_idx, (header, _) in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            # Write data
            for row_idx, program in enumerate(programs, 2):
                for col_idx, (_, field_name) in enumerate(columns, 1):
                    value = program.get(field_name, "")
                    if isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(
                        horizontal="left", vertical="top", wrap_text=True
                    )

            # Adjust column widths
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 20
            ws.column_dimensions["F"].width = 25
            ws.column_dimensions["G"].width = 20
            ws.column_dimensions["H"].width = 15
            ws.column_dimensions["I"].width = 30
            ws.column_dimensions["J"].width = 15
            ws.column_dimensions["K"].width = 15
            ws.column_dimensions["L"].width = 10
            ws.column_dimensions["M"].width = 30
            ws.column_dimensions["N"].width = 12
            ws.column_dimensions["O"].width = 20

            # Freeze header row
            ws.freeze_panes = "A2"

            # Save workbook
            wb.save(filepath)
            logger.info(f"[OK] Exported {len(programs)} programs to Excel: {filepath}")
            return True

        except Exception as e:
            logger.error(f"[ERROR] Failed to export to Excel: {e}")
            return False

    def export_to_json(
        self, programs: List[Dict[str, Any]], filename: str = None
    ) -> bool:
        """Export programs to JSON file"""
        try:
            if filename is None:
                filename = f"programs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

            filepath = os.path.join(self.export_dir, filename)

            if not programs:
                logger.warning("[WARNING] No programs to export to JSON")
                return False

            # Convert datetime objects to strings
            programs_serializable = []
            for program in programs:
                program_copy = program.copy()
                if "extracted_at" in program_copy and isinstance(
                    program_copy["extracted_at"], datetime
                ):
                    program_copy["extracted_at"] = program_copy[
                        "extracted_at"
                    ].isoformat()
                programs_serializable.append(program_copy)

            # Write JSON
            with open(filepath, "w", encoding="utf-8") as jsonfile:
                json.dump(programs_serializable, jsonfile, indent=2, ensure_ascii=False)

            logger.info(f"[OK] Exported {len(programs)} programs to JSON: {filepath}")
            return True

        except Exception as e:
            logger.error(f"[ERROR] Failed to export to JSON: {e}")
            return False

    def export_to_pandas_excel(
        self, programs: List[Dict[str, Any]], filename: str = None
    ) -> bool:
        """Export programs to Excel using pandas (requires pandas)"""
        try:
            if not PANDAS_AVAILABLE:
                logger.warning(
                    "[WARNING] pandas not installed. Use export_to_excel() instead."
                )
                return False

            if filename is None:
                filename = (
                    f"programs_pandas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )

            filepath = os.path.join(self.export_dir, filename)

            if not programs:
                logger.warning("[WARNING] No programs to export")
                return False

            # Create DataFrame
            df = pd.DataFrame(programs)

            # Reorder columns if they exist
            preferred_columns = [
                "university_name",
                "program_name",
                "degree_type",
                "duration",
                "tuition_fees",
                "admission_requirements",
                "language_requirements",
                "application_deadline",
                "program_url",
                "country",
                "city",
                "ranking",
                "description",
                "confidence_score",
                "extracted_at",
            ]

            existing_columns = [col for col in preferred_columns if col in df.columns]
            df = df[existing_columns]

            # Export to Excel with formatting
            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Programs", index=False)

                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets["Programs"]

                # Apply formatting
                header_fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                header_font = Font(bold=True, color="FFFFFF")

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)

                # Adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            logger.info(
                f"[OK] Exported {len(programs)} programs to Excel (pandas): {filepath}"
            )
            return True

        except Exception as e:
            logger.error(f"[ERROR] Failed to export to Excel (pandas): {e}")
            return False

    def save_to_mongodb(self, programs: List[Dict[str, Any]]) -> int:
        """Save programs to MongoDB database"""
        try:
            if not self.collection:
                logger.error("[ERROR] MongoDB not connected")
                return 0

            saved_count = 0
            for program in programs:
                try:
                    # Convert datetime to string if needed
                    program_copy = program.copy()
                    if "extracted_at" in program_copy and isinstance(
                        program_copy["extracted_at"], datetime
                    ):
                        program_copy["extracted_at"] = program_copy[
                            "extracted_at"
                        ].isoformat()

                    # Upsert based on university_name and program_name
                    result = self.collection.update_one(
                        {
                            "university_name": program_copy["university_name"],
                            "program_name": program_copy["program_name"],
                        },
                        {"$set": program_copy},
                        upsert=True,
                    )

                    if result.upserted_id or result.modified_count > 0:
                        saved_count += 1

                except Exception as e:
                    logger.error(
                        f"[ERROR] Failed to save program {program.get('program_name', 'Unknown')}: {e}"
                    )
                    continue

            logger.info(f"[OK] Saved {saved_count} programs to MongoDB")
            return saved_count

        except Exception as e:
            logger.error(f"[ERROR] Failed to save to MongoDB: {e}")
            return 0

    def export_all_formats(self, programs: List[Dict[str, Any]]) -> Dict[str, bool]:
        """Export to all available formats"""
        results = {
            "csv": self.export_to_csv(programs),
            "excel": self.export_to_excel(programs),
            "json": self.export_to_json(programs),
            "mongodb": self.save_to_mongodb(programs) > 0,
        }

        if PANDAS_AVAILABLE:
            results["excel_pandas"] = self.export_to_pandas_excel(programs)

        return results

    def get_exported_files(self) -> List[str]:
        """Get list of exported files"""
        try:
            files = os.listdir(self.export_dir)
            return sorted(files, reverse=True)
        except Exception as e:
            logger.error(f"[ERROR] Failed to list exported files: {e}")
            return []

    def close(self):
        """Close MongoDB connection"""
        if self.mongo_client:
            self.mongo_client.close()
            logger.info("[OK] MongoDB connection closed")


def main():
    """Example usage of data exporter"""
    # Sample programs data
    sample_programs = [
        {
            "university_name": "Harvard University",
            "program_name": "Computer Science",
            "degree_type": "Master",
            "duration": "2 years",
            "tuition_fees": "$60,000 per year",
            "admission_requirements": "Bachelor's degree",
            "language_requirements": "English",
            "application_deadline": "December 15",
            "program_url": "https://www.harvard.edu/academics/",
            "country": "United States",
            "city": "Cambridge",
            "ranking": "1",
            "description": "Advanced computer science program",
            "confidence_score": 0.95,
            "extracted_at": datetime.now(),
        }
    ]

    # Initialize exporter
    exporter = UniversityDataExporter()

    # Connect to MongoDB
    if exporter.connect_mongodb():
        # Export to all formats
        print("\n" + "=" * 80)
        print("EXPORTING DATA TO MULTIPLE FORMATS")
        print("=" * 80)

        results = exporter.export_all_formats(sample_programs)

        print("\nExport Results:")
        for format_name, success in results.items():
            status = "✅ SUCCESS" if success else "❌ FAILED"
            print(f"  {format_name.upper():20} {status}")

        # List exported files
        files = exporter.get_exported_files()
        if files:
            print(f"\nExported files ({len(files)} total):")
            for file in files[:10]:  # Show first 10
                print(f"  • {file}")

        # Close connection
        exporter.close()
    else:
        print("[ERROR] Failed to connect to MongoDB")


if __name__ == "__main__":
    main()
