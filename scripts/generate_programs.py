#!/usr/bin/env python3
"""
Phase 1: Generate university program records from QS Rankings data.

Transforms 1,504 QS-ranked universities into structured UniversityProgram
records using country/tier-based defaults for tuition, GPA, IELTS, etc.

Output: MongoDB + CSV export

Fixes applied:
  P1: _safe_float() returns NaN for None (not 0.0)
  P2: Pydantic schema validation on every generated record
  P5: Column name resolution instead of hardcoded indices
"""

import csv
import hashlib
import json
import logging
import math
import random
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

# Add project root to path
sys.path.insert(0, str(Path(__file__).parents[1]))

import openpyxl
from pydantic import BaseModel, Field, field_validator

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


# ─── P2 FIX: PYDANTIC SCHEMA FOR GENERATED RECORDS ─────────────────────────


class ProgramRecord(BaseModel):
    """Validates a generated program record before it enters the dataset."""

    university_id: str
    university_name: str = Field(min_length=1)
    program_name: str = Field(min_length=1)
    degree_type: str
    degree_level: str
    program_category: str
    country: str
    city: str
    gpa_requirement_min: float = Field(ge=0.0, le=4.0)
    ielts_min: Optional[float] = Field(default=None, ge=0.0, le=9.0)
    toefl_min: Optional[int] = Field(default=None, ge=0, le=120)
    duration_years: float = Field(gt=0, le=10.0)
    tuition_domestic: int = Field(ge=0)
    tuition_international: int = Field(ge=0)
    tuition_currency: str = Field(min_length=2, max_length=4)
    cost_of_living: int = Field(ge=0)
    application_fee: int = Field(ge=0, le=500)
    qs_world_ranking: int = Field(ge=1)
    qs_overall_score: float = Field(ge=0.0)
    university_tier: str

    @field_validator("gpa_requirement_min")
    @classmethod
    def gpa_must_be_reasonable(cls, v: float) -> float:
        if v > 4.0 or v < 0.0:
            raise ValueError(f"GPA {v} outside valid range [0.0, 4.0]")
        return v

    @field_validator("university_tier")
    @classmethod
    def tier_must_be_valid(cls, v: str) -> str:
        if v not in {"top", "good", "standard"}:
            raise ValueError(f"Invalid tier: {v}")
        return v


# ─── COUNTRY-BASED DEFAULTS ──────────────────────────────────────────────────

COUNTRY_TUITION = {
    # (domestic_per_year_usd, international_per_year_usd, currency)
    "United States of America": (15000, 45000, "USD"),
    "United Kingdom": (12000, 30000, "GBP"),
    "Canada": (8000, 30000, "CAD"),
    "Australia": (10000, 35000, "AUD"),
    "Germany": (500, 1500, "EUR"),
    "France": (300, 5000, "EUR"),
    "Netherlands": (2200, 15000, "EUR"),
    "Switzerland": (1500, 3000, "CHF"),
    "Sweden": (0, 15000, "SEK"),
    "Denmark": (0, 15000, "DKK"),
    "Norway": (0, 0, "NOK"),
    "Finland": (0, 15000, "EUR"),
    "Japan": (5000, 10000, "JPY"),
    "South Korea": (5000, 10000, "KRW"),
    "China (Mainland)": (3000, 6000, "CNY"),
    "Hong Kong SAR": (6000, 18000, "HKD"),
    "Singapore": (8000, 20000, "SGD"),
    "India": (1000, 5000, "INR"),
    "Malaysia": (2000, 8000, "MYR"),
    "Ireland": (5000, 18000, "EUR"),
    "Italy": (2000, 5000, "EUR"),
    "Spain": (1500, 8000, "EUR"),
    "Belgium": (1000, 5000, "EUR"),
    "Austria": (800, 1600, "EUR"),
    "New Zealand": (6000, 25000, "NZD"),
    "Russia": (2000, 5000, "RUB"),
    "Brazil": (0, 3000, "BRL"),
    "Mexico": (1000, 5000, "MXN"),
    "Saudi Arabia": (0, 0, "SAR"),
    "United Arab Emirates": (10000, 20000, "AED"),
    "Turkey": (500, 3000, "TRY"),
    "Poland": (0, 4000, "PLN"),
    "Czech Republic": (0, 5000, "CZK"),
    "Portugal": (1000, 5000, "EUR"),
    "Taiwan": (3000, 6000, "TWD"),
    "Thailand": (1500, 5000, "THB"),
    "Indonesia": (1000, 4000, "IDR"),
    "Pakistan": (500, 3000, "PKR"),
    "Egypt": (500, 5000, "EGP"),
    "South Africa": (3000, 8000, "ZAR"),
    "Argentina": (0, 3000, "ARS"),
    "Chile": (5000, 10000, "CLP"),
    "Colombia": (2000, 6000, "COP"),
    "Philippines": (1000, 4000, "PHP"),
    "Israel": (5000, 12000, "ILS"),
}

# Default for unlisted countries
DEFAULT_TUITION = (3000, 10000, "USD")

# Country → typical IELTS requirement
COUNTRY_IELTS = {
    "United States of America": 6.5,
    "United Kingdom": 6.5,
    "Canada": 6.5,
    "Australia": 6.5,
    "New Zealand": 6.5,
    "Ireland": 6.5,
    "Singapore": 6.5,
    "Hong Kong SAR": 6.5,
    "Netherlands": 6.5,
    "Sweden": 6.5,
    "Denmark": 6.5,
    "Norway": 6.5,
    "Finland": 6.0,
    "Germany": 6.0,
    "France": 6.0,
    "Switzerland": 6.5,
    "Japan": 6.0,
    "South Korea": 6.0,
    "Malaysia": 6.0,
}
DEFAULT_IELTS = 6.0

# Country → typical TOEFL iBT
COUNTRY_TOEFL = {
    "United States of America": 90,
    "United Kingdom": 90,
    "Canada": 90,
    "Australia": 85,
    "New Zealand": 85,
    "Ireland": 90,
    "Singapore": 90,
    "Hong Kong SAR": 85,
    "Netherlands": 85,
    "Germany": 80,
    "France": 80,
    "Switzerland": 85,
    "Sweden": 85,
}
DEFAULT_TOEFL = 80

# Tier → GPA ranges
TIER_GPA = {
    "top": (3.5, 3.7),  # rank 1-50
    "good": (3.0, 3.4),  # rank 51-200
    "standard": (2.5, 3.0),  # rank 201+
}

# Cost of living (annual USD) by country
COUNTRY_COL = {
    "United States of America": 18000,
    "United Kingdom": 15000,
    "Canada": 14000,
    "Australia": 16000,
    "Switzerland": 25000,
    "Singapore": 15000,
    "Hong Kong SAR": 16000,
    "Japan": 12000,
    "Germany": 11000,
    "France": 12000,
    "Netherlands": 13000,
    "Sweden": 12000,
    "Norway": 15000,
    "Denmark": 14000,
    "South Korea": 10000,
    "Ireland": 14000,
    "New Zealand": 13000,
    "Italy": 11000,
    "Spain": 10000,
    "Belgium": 12000,
}
DEFAULT_COL = 8000

# ─── PROGRAM TEMPLATES ──────────────────────────────────────────────────────

PROGRAM_TEMPLATES = {
    "Computer Science": {
        "programs": [
            ("Computer Science", "MSc"),
            ("Artificial Intelligence", "MSc"),
            ("Data Science", "MSc"),
            ("Computer Science", "BSc"),
            ("Software Engineering", "MSc"),
            ("Computer Science", "PhD"),
        ],
        "category": "Computer Science",
        "specializations": [
            "Machine Learning",
            "Cybersecurity",
            "Software Engineering",
            "Data Analytics",
            "Cloud Computing",
        ],
        "careers": [
            "Software Engineer",
            "Data Scientist",
            "ML Engineer",
            "Systems Architect",
            "DevOps Engineer",
        ],
    },
    "Business": {
        "programs": [
            ("Business Administration", "MBA"),
            ("Finance", "MSc"),
            ("Marketing", "MSc"),
            ("Management", "BSc"),
            ("International Business", "MSc"),
            ("Business Administration", "PhD"),
        ],
        "category": "Business",
        "specializations": [
            "Strategic Management",
            "Digital Marketing",
            "Financial Analysis",
            "Entrepreneurship",
            "Supply Chain",
        ],
        "careers": [
            "Business Analyst",
            "Financial Analyst",
            "Marketing Manager",
            "Consultant",
            "Product Manager",
        ],
    },
    "Engineering": {
        "programs": [
            ("Mechanical Engineering", "MSc"),
            ("Electrical Engineering", "MSc"),
            ("Civil Engineering", "MSc"),
            ("Chemical Engineering", "BSc"),
            ("Aerospace Engineering", "MSc"),
            ("Engineering", "PhD"),
        ],
        "category": "Engineering",
        "specializations": [
            "Robotics",
            "Renewable Energy",
            "Structural Design",
            "Nanotechnology",
            "Control Systems",
        ],
        "careers": [
            "Design Engineer",
            "Project Engineer",
            "R&D Engineer",
            "Systems Engineer",
            "Technical Lead",
        ],
    },
    "Medical": {
        "programs": [
            ("Medicine", "MD"),
            ("Public Health", "MPH"),
            ("Biomedical Sciences", "MSc"),
            ("Nursing", "BSc"),
            ("Pharmacy", "BSc"),
            ("Biomedical Sciences", "PhD"),
        ],
        "category": "Medical",
        "specializations": [
            "Epidemiology",
            "Clinical Research",
            "Health Informatics",
            "Neuroscience",
            "Genetics",
        ],
        "careers": [
            "Physician",
            "Research Scientist",
            "Public Health Officer",
            "Clinical Analyst",
            "Healthcare Manager",
        ],
    },
    "Law": {
        "programs": [
            ("Law", "LLM"),
            ("International Law", "LLM"),
            ("Law", "LLB"),
            ("Law", "PhD"),
        ],
        "category": "Law",
        "specializations": [
            "Corporate Law",
            "Human Rights",
            "International Trade",
            "Intellectual Property",
            "Criminal Law",
        ],
        "careers": [
            "Lawyer",
            "Legal Advisor",
            "Compliance Officer",
            "Judge",
            "Policy Analyst",
        ],
    },
    "Sciences": {
        "programs": [
            ("Physics", "MSc"),
            ("Mathematics", "MSc"),
            ("Chemistry", "BSc"),
            ("Biology", "BSc"),
            ("Environmental Science", "MSc"),
            ("Physics", "PhD"),
        ],
        "category": "Sciences",
        "specializations": [
            "Quantum Physics",
            "Applied Mathematics",
            "Organic Chemistry",
            "Molecular Biology",
            "Climate Science",
        ],
        "careers": [
            "Research Scientist",
            "Lab Manager",
            "Data Analyst",
            "Environmental Consultant",
            "Professor",
        ],
    },
    "Arts & Humanities": {
        "programs": [
            ("English Literature", "MA"),
            ("History", "MA"),
            ("Philosophy", "BA"),
            ("Linguistics", "MSc"),
            ("History", "PhD"),
        ],
        "category": "Arts & Humanities",
        "specializations": [
            "Modern Literature",
            "Ancient History",
            "Ethics",
            "Sociolinguistics",
            "Comparative Studies",
        ],
        "careers": [
            "Writer",
            "Educator",
            "Archivist",
            "Translator",
            "Cultural Analyst",
        ],
    },
    "Education": {
        "programs": [
            ("Education", "MEd"),
            ("TESOL", "MA"),
            ("Educational Psychology", "MSc"),
            ("Education", "PhD"),
        ],
        "category": "Education",
        "specializations": [
            "Curriculum Design",
            "Educational Technology",
            "Special Education",
            "Assessment",
            "Leadership",
        ],
        "careers": [
            "Teacher",
            "School Administrator",
            "Curriculum Developer",
            "EdTech Specialist",
            "Counselor",
        ],
    },
}

# Degree type mapping
DEGREE_ABBREV_TO_FULL = {
    "BSc": "Bachelor of Science",
    "BA": "Bachelor of Arts",
    "MSc": "Master of Science",
    "MA": "Master of Arts",
    "PhD": "Doctor of Philosophy",
    "MBA": "Master of Business Administration",
    "MEd": "Master of Education",
    "MPH": "Master of Public Health",
    "LLM": "Master of Laws",
    "LLB": "Bachelor of Laws",
    "MD": "Doctor of Medicine",
    "PhD": "Doctor of Philosophy",
}

DEGREE_ABBREV_TO_LEVEL = {
    "BSc": "Undergraduate",
    "BA": "Undergraduate",
    "LLB": "Undergraduate",
    "MSc": "Masters",
    "MA": "Masters",
    "MBA": "Masters",
    "MEd": "Masters",
    "MPH": "Masters",
    "LLM": "Masters",
    "PhD": "PhD",
    "MD": "Masters",
    "PhD": "PhD",
}

DEGREE_DURATION = {
    "BSc": 3.0,
    "BA": 3.0,
    "LLB": 3.0,
    "MSc": 1.5,
    "MA": 1.5,
    "MBA": 2.0,
    "MEd": 1.0,
    "MPH": 1.5,
    "LLM": 1.0,
    "MD": 5.0,
    "PhD": 3.5,
}


# ─── QS EXCEL PARSER ────────────────────────────────────────────────────────

# P5 FIX: Mapping from our internal key to possible header names in the Excel.
# Column matching is case-insensitive and uses substring matching.
# This replaces hardcoded row[1], row[3], etc.
QS_COLUMN_MAP = {
    "rank": ["rank", "2026 rank", "ranking"],
    "name": ["institution name", "university", "institution"],
    "country": ["location", "country"],
    "region": ["region"],
    "size": ["size"],
    "focus": ["focus"],
    "research": ["res.", "research"],
    "status": ["status"],
    "academic_reputation": ["academic reputation"],
    "employer_reputation": ["employer reputation"],
    "faculty_student_ratio": ["faculty student ratio", "faculty student"],
    "citations": ["citations per faculty", "citations"],
    "intl_faculty": ["international faculty ratio", "international faculty"],
    "intl_students": ["international students ratio", "international students"],
    "employment_outcomes": ["employment outcomes"],
    "sustainability": ["sustainability"],
    "overall_score": ["overall score", "overall"],
}


def _build_column_index(ws) -> dict:
    """
    P5 FIX: Read header rows and build a {our_key → column_index} mapping.

    Scans the first 3 rows for header text, matching against QS_COLUMN_MAP.
    Returns a dict mapping our internal keys to 0-based column indices.
    """
    # Collect all text from the first 3 rows
    header_texts = {}  # col_idx → combined header text
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True)):
        for col_idx, cell_val in enumerate(row):
            if cell_val is not None:
                text = str(cell_val).strip().lower()
                existing = header_texts.get(col_idx, "")
                header_texts[col_idx] = f"{existing} {text}".strip()

    # Match our keys against the combined header text
    col_map = {}
    for our_key, candidates in QS_COLUMN_MAP.items():
        for col_idx, header_text in header_texts.items():
            for candidate in candidates:
                if candidate.lower() in header_text:
                    col_map[our_key] = col_idx
                    break
            if our_key in col_map:
                break

    logger.info(
        f"Column mapping resolved: {len(col_map)}/{len(QS_COLUMN_MAP)} columns matched"
    )
    for key, idx in sorted(col_map.items(), key=lambda x: x[1]):
        logger.debug(f"  {key} → column {idx} (header: '{header_texts.get(idx, '?')}')")

    # Verify required columns
    required = {"rank", "name"}
    missing = required - set(col_map.keys())
    if missing:
        raise ValueError(
            f"Required columns not found in Excel headers: {missing}. "
            f"Available header texts: {dict(list(header_texts.items())[:10])}"
        )

    return col_map


def _get_cell(row: tuple, col_map: dict, key: str, default=None):
    """Safely get a cell value by mapped column name."""
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return default
    val = row[idx]
    return val if val is not None else default


def parse_qs_excel(filepath: str) -> list:
    """Parse QS Rankings Excel into list of university dicts.

    P5 FIX: Uses column name resolution instead of hardcoded indices.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Build column mapping from header rows
    col_map = _build_column_index(ws)

    universities = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        # Skip empty rows
        name_val = _get_cell(row, col_map, "name")
        if not name_val:
            continue

        rank = _get_cell(row, col_map, "rank")
        # Handle string ranks like "1001-1200"
        if isinstance(rank, str) and "-" in rank:
            rank = int(rank.split("-")[0])
        elif isinstance(rank, str) and "+" in rank:
            rank = int(rank.replace("+", ""))
        elif rank is None:
            continue
        else:
            try:
                rank = int(rank)
            except (ValueError, TypeError):
                continue

        uni = {
            "rank": rank,
            "name": str(name_val).strip(),
            "country": str(_get_cell(row, col_map, "country", "Unknown")).strip(),
            "region": str(_get_cell(row, col_map, "region", "Unknown")).strip(),
            "size": str(_get_cell(row, col_map, "size", "M")).strip(),
            "focus": str(_get_cell(row, col_map, "focus", "CO")).strip(),
            "research": str(_get_cell(row, col_map, "research", "M")).strip(),
            "status": str(_get_cell(row, col_map, "status", "Unknown")).strip(),
            "academic_reputation": _safe_float(
                _get_cell(row, col_map, "academic_reputation")
            ),
            "employer_reputation": _safe_float(
                _get_cell(row, col_map, "employer_reputation")
            ),
            "faculty_student_ratio": _safe_float(
                _get_cell(row, col_map, "faculty_student_ratio")
            ),
            "citations": _safe_float(_get_cell(row, col_map, "citations")),
            "intl_faculty": _safe_float(_get_cell(row, col_map, "intl_faculty")),
            "intl_students": _safe_float(_get_cell(row, col_map, "intl_students")),
            "employment_outcomes": _safe_float(
                _get_cell(row, col_map, "employment_outcomes")
            ),
            "sustainability": _safe_float(_get_cell(row, col_map, "sustainability")),
            "overall_score": _safe_float(_get_cell(row, col_map, "overall_score")),
        }
        universities.append(uni)

    wb.close()
    logger.info(f"Parsed {len(universities)} universities from QS Rankings")
    return universities


def _safe_float(val) -> float:
    """P1 FIX: Return NaN for None/invalid values instead of 0.0.

    Returning 0.0 for missing data is dangerous because it is
    indistinguishable from a legitimate score of 0.0. NaN allows
    downstream imputation (median/mean) to handle missing values
    correctly.
    """
    if val is None:
        return float("nan")
    try:
        result = float(val)
        return result
    except (ValueError, TypeError):
        return float("nan")


# ─── PROGRAM GENERATOR ──────────────────────────────────────────────────────


def get_tier(rank: int) -> str:
    if rank <= 50:
        return "top"
    elif rank <= 200:
        return "good"
    return "standard"


def get_city(name: str, country: str) -> str:
    """Extract city from university name or use country capital."""
    # Common patterns
    city_patterns = {
        "London": ["London", "UCL", "Imperial", "King's College London", "LSE"],
        "Cambridge": ["Cambridge"],
        "Oxford": ["Oxford"],
        "Edinburgh": ["Edinburgh"],
        "Manchester": ["Manchester"],
        "Boston": ["MIT", "Harvard", "Boston"],
        "Stanford": ["Stanford"],
        "New York": ["Columbia", "NYU", "New York"],
        "Chicago": ["Chicago"],
        "Toronto": ["Toronto"],
        "Melbourne": ["Melbourne"],
        "Sydney": ["Sydney"],
        "Zurich": ["ETH Zurich", "Zurich"],
        "Munich": ["Munich", "München", "TUM", "LMU"],
        "Berlin": ["Berlin", "Humboldt", "Freie"],
        "Paris": ["Paris", "PSL", "Sorbonne", "Polytechnique"],
        "Amsterdam": ["Amsterdam"],
        "Tokyo": ["Tokyo"],
        "Beijing": ["Peking", "Tsinghua", "Beijing"],
        "Shanghai": ["Shanghai", "Fudan", "Jiao Tong"],
        "Singapore": ["Singapore", "NUS", "NTU"],
        "Hong Kong": ["Hong Kong"],
        "Seoul": ["Seoul", "Korea", "KAIST"],
    }
    for city, patterns in city_patterns.items():
        if any(p.lower() in name.lower() for p in patterns):
            return city

    # Country capitals as fallback
    capitals = {
        "United States of America": "Washington, D.C.",
        "United Kingdom": "London",
        "Canada": "Ottawa",
        "Australia": "Canberra",
        "Germany": "Berlin",
        "France": "Paris",
        "Japan": "Tokyo",
        "China (Mainland)": "Beijing",
        "India": "New Delhi",
        "South Korea": "Seoul",
        "Netherlands": "Amsterdam",
        "Switzerland": "Bern",
        "Sweden": "Stockholm",
        "Italy": "Rome",
        "Spain": "Madrid",
        "Brazil": "Brasília",
        "Russia": "Moscow",
    }
    return capitals.get(country, country)


def generate_programs_for_university(uni: dict) -> list:
    """Generate program records for a single university.

    P2 FIX: Each record is validated via ProgramRecord before being accepted.
    Invalid records are logged and skipped.
    """
    rank = uni["rank"]
    tier = get_tier(rank)
    country = uni["country"]
    city = get_city(uni["name"], country)

    # Get tuition defaults
    tuition_dom, tuition_intl, currency = COUNTRY_TUITION.get(country, DEFAULT_TUITION)

    # Adjust tuition by tier
    tier_multiplier = {"top": 1.3, "good": 1.0, "standard": 0.8}[tier]
    tuition_dom = int(tuition_dom * tier_multiplier)
    tuition_intl = int(tuition_intl * tier_multiplier)

    # NOISE: Add ±15% jitter to break deterministic country→tier→cost mapping.
    # Without noise the model learns a lookup table, not generalizable patterns.
    def _jitter(value, pct=0.15):
        return int(value * (1 + random.uniform(-pct, pct)))

    # Import currency normalizer for USD columns
    try:
        from pipeline.currency import to_usd
    except ImportError:
        import sys as _sys

        _sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
        from pipeline.currency import to_usd

    # GPA range
    gpa_min, gpa_max = TIER_GPA[tier]

    # Language requirements
    ielts = COUNTRY_IELTS.get(country, DEFAULT_IELTS)
    toefl = COUNTRY_TOEFL.get(country, DEFAULT_TOEFL)

    # Cost of living
    col = COUNTRY_COL.get(country, DEFAULT_COL)

    # Select which program categories to generate
    # Top universities get all categories, lower-ranked get fewer
    categories = list(PROGRAM_TEMPLATES.keys())
    if rank > 500:
        # Standard: 3-5 categories
        random.seed(hash(uni["name"]))
        num_categories = random.randint(3, 5)
        categories = random.sample(categories, min(num_categories, len(categories)))
    elif rank > 200:
        # Good: 5-7 categories
        random.seed(hash(uni["name"]))
        num_categories = random.randint(5, 7)
        categories = random.sample(categories, min(num_categories, len(categories)))

    # Select programs per category (top unis get more programs)
    programs_per_category = 3 if rank <= 100 else 2 if rank <= 500 else 1

    # PhD programs for research-active universities (rank <= 300)
    include_phd = rank <= 300

    programs = []
    from src.models.university import generate_university_id

    uni_id = generate_university_id(uni["name"])

    # P1 FIX: safe NaN-aware access for QS overall score
    qs_overall = uni.get("overall_score", float("nan"))
    if isinstance(qs_overall, float) and math.isnan(qs_overall):
        qs_overall = 0.0  # for CSV export only — NOT used as training feature

    for cat_name in categories:
        template = PROGRAM_TEMPLATES[cat_name]
        # Separate non-PhD and PhD programs
        non_phd = [(n, d) for n, d in template["programs"] if d != "PhD"]
        phd_progs = [(n, d) for n, d in template["programs"] if d == "PhD"]

        selected_programs = non_phd[:programs_per_category]
        # Always add PhD for research universities
        if include_phd and phd_progs:
            selected_programs.extend(phd_progs)

        for prog_name, degree_abbrev in selected_programs:
            # Slight GPA jitter per program
            random.seed(hash(f"{uni['name']}-{prog_name}-{degree_abbrev}"))
            gpa = round(random.uniform(gpa_min, gpa_max), 1)
            duration = DEGREE_DURATION.get(degree_abbrev, 1.5)

            # NOISE: jitter per-program so identical universities have variance
            prog_tuition_dom = _jitter(tuition_dom)
            prog_tuition_intl = _jitter(tuition_intl)
            prog_col = _jitter(col)
            # USD normalized versions for training
            prog_tuition_dom_usd = to_usd(prog_tuition_dom, currency)
            prog_tuition_intl_usd = to_usd(prog_tuition_intl, currency)
            prog_col_usd = float(prog_col)  # COL is already in USD

            # US bachelor programs are 4 years, UK are 3
            if (
                degree_abbrev in ("BSc", "BA", "LLB")
                and country == "United States of America"
            ):
                duration = 4.0

            # ── P2 FIX: Validate record before accepting ─────────────
            try:
                validated = ProgramRecord(
                    university_id=uni_id,
                    university_name=uni["name"],
                    program_name=prog_name,
                    degree_type=DEGREE_ABBREV_TO_FULL.get(
                        degree_abbrev, "Master of Science"
                    ),
                    degree_level=DEGREE_ABBREV_TO_LEVEL.get(degree_abbrev, "Masters"),
                    program_category=template["category"],
                    country=country,
                    city=city,
                    gpa_requirement_min=gpa,
                    ielts_min=ielts,
                    toefl_min=int(toefl),
                    duration_years=duration,
                    tuition_domestic=prog_tuition_dom,
                    tuition_international=prog_tuition_intl,
                    tuition_currency=currency,
                    cost_of_living=prog_col,
                    application_fee=75
                    if country == "United States of America"
                    else 50
                    if country == "United Kingdom"
                    else 30,
                    qs_world_ranking=rank,
                    qs_overall_score=qs_overall,
                    university_tier=tier,
                )
            except Exception as e:
                logger.warning(
                    f"SKIPPING invalid record: {uni['name']} / {prog_name} "
                    f"({degree_abbrev}): {e}"
                )
                continue

            program = {
                "university_id": uni_id,
                "university_name": uni["name"],
                "program_name": f"{prog_name}",
                "degree_type": DEGREE_ABBREV_TO_FULL.get(
                    degree_abbrev, "Master of Science"
                ),
                "degree_level": DEGREE_ABBREV_TO_LEVEL.get(degree_abbrev, "Masters"),
                "program_category": template["category"],
                "country": country,
                "city": city,
                "gpa_requirement_min": gpa,
                "language_requirements": {
                    "ielts_min": ielts,
                    "toefl_min": toefl,
                },
                "prerequisites": [],
                "duration_years": duration,
                "program_description": f"{prog_name} program at {uni['name']}, ranked #{rank} globally by QS.",
                "specializations": random.sample(
                    template["specializations"],
                    min(3, len(template["specializations"])),
                ),
                "faculty_research_interests": [],
                "tuition_fees": {
                    "domestic_per_year": prog_tuition_dom,
                    "international_per_year": prog_tuition_intl,
                    "currency": currency,
                },
                "scholarships_available": [],
                "cost_of_living_estimate": prog_col,
                # USD-normalized columns for ML training
                "tuition_domestic_usd": prog_tuition_dom_usd,
                "tuition_international_usd": prog_tuition_intl_usd,
                "cost_of_living_usd": prog_col_usd,
                "application_deadlines": {
                    "fall_deadline": None,
                    "spring_deadline": None,
                    "rolling_admission": rank > 300,
                },
                "application_fee": 75
                if country == "United States of America"
                else 50
                if country == "United Kingdom"
                else 30,
                "rankings": {
                    "qs_world_ranking": rank,
                },
                "career_outcomes": random.sample(
                    template["careers"], min(3, len(template["careers"]))
                ),
                "research_opportunities": [],
                "campus_facilities": [],
                "source_url": None,
                "confidence_score": 0.7,  # Synthetic data confidence
                "last_updated": datetime.now(timezone.utc).isoformat(),
                "data_completeness": 0.65,
                "data_source": "qs_rankings_generated",
                # Extra fields for ML
                "_qs_overall_score": uni["overall_score"],
                "_qs_academic_reputation": uni["academic_reputation"],
                "_qs_employer_reputation": uni["employer_reputation"],
                "_qs_faculty_student_ratio": uni["faculty_student_ratio"],
                "_qs_citations": uni["citations"],
                "_qs_intl_students": uni["intl_students"],
                "_qs_employment_outcomes": uni["employment_outcomes"],
                "_qs_sustainability": uni["sustainability"],
                "_university_size": uni["size"],
                "_university_focus": uni["focus"],
                "_university_research": uni["research"],
                "_university_status": uni["status"],
                "_university_tier": tier,
            }
            programs.append(program)

    return programs


# ─── EXPORT ─────────────────────────────────────────────────────────────────


def export_csv(programs: list, filepath: str):
    """Export programs to flat CSV for ML training."""
    if not programs:
        return

    # Flatten nested fields
    flat_records = []
    for p in programs:
        flat = {
            "university_id": p["university_id"],
            "university_name": p["university_name"],
            "program_name": p["program_name"],
            "degree_type": p["degree_type"],
            "degree_level": p["degree_level"],
            "program_category": p["program_category"],
            "country": p["country"],
            "city": p["city"],
            "gpa_requirement_min": p["gpa_requirement_min"],
            "ielts_min": p["language_requirements"].get("ielts_min"),
            "toefl_min": p["language_requirements"].get("toefl_min"),
            "duration_years": p["duration_years"],
            "tuition_domestic": p["tuition_fees"].get("domestic_per_year"),
            "tuition_international": p["tuition_fees"].get("international_per_year"),
            "tuition_currency": p["tuition_fees"].get("currency"),
            "cost_of_living": p["cost_of_living_estimate"],
            "application_fee": p["application_fee"],
            "qs_world_ranking": p["rankings"].get("qs_world_ranking"),
            "rolling_admission": p["application_deadlines"].get("rolling_admission"),
            "confidence_score": p["confidence_score"],
            "data_completeness": p["data_completeness"],
            "specializations": "; ".join(p.get("specializations", [])),
            "career_outcomes": "; ".join(p.get("career_outcomes", [])),
            # QS indicator scores
            "qs_overall_score": p.get("_qs_overall_score", 0),
            "qs_academic_reputation": p.get("_qs_academic_reputation", 0),
            "qs_employer_reputation": p.get("_qs_employer_reputation", 0),
            "qs_faculty_student_ratio": p.get("_qs_faculty_student_ratio", 0),
            "qs_citations": p.get("_qs_citations", 0),
            "qs_intl_students": p.get("_qs_intl_students", 0),
            "qs_employment_outcomes": p.get("_qs_employment_outcomes", 0),
            "qs_sustainability": p.get("_qs_sustainability", 0),
            "university_size": p.get("_university_size", ""),
            "university_focus": p.get("_university_focus", ""),
            "university_research": p.get("_university_research", ""),
            "university_status": p.get("_university_status", ""),
            "university_tier": p.get("_university_tier", ""),
        }
        flat_records.append(flat)

    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=flat_records[0].keys())
        writer.writeheader()
        writer.writerows(flat_records)

    logger.info(f"Exported {len(flat_records)} records to {filepath}")


def export_json(programs: list, filepath: str):
    """Export programs to JSON."""
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    # Remove internal fields
    clean = []
    for p in programs:
        c = {k: v for k, v in p.items() if not k.startswith("_")}
        clean.append(c)

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(clean, f, indent=2, default=str)

    logger.info(f"Exported {len(clean)} records to {filepath}")


# ─── MAIN ───────────────────────────────────────────────────────────────────


def main():
    xlsx_path = "data/2026 QS World University Rankings 1.3 (For qs.com).xlsx"

    logger.info("=" * 60)
    logger.info("PHASE 1: QS Rankings → Program Records")
    logger.info("=" * 60)

    # Parse QS data
    universities = parse_qs_excel(xlsx_path)
    logger.info(f"Loaded {len(universities)} universities")

    # Generate programs
    all_programs = []
    for uni in universities:
        programs = generate_programs_for_university(uni)
        all_programs.extend(programs)

    logger.info(f"Generated {len(all_programs)} program records")

    # Stats
    countries = set(p["country"] for p in all_programs)
    categories = set(p["program_category"] for p in all_programs)
    levels = set(p["degree_level"] for p in all_programs)
    unis = set(p["university_name"] for p in all_programs)

    logger.info(
        f"Coverage: {len(unis)} universities, {len(countries)} countries, "
        f"{len(categories)} categories, {len(levels)} degree levels"
    )

    # Distribution
    logger.info("\n--- Category Distribution ---")
    from collections import Counter

    cat_dist = Counter(p["program_category"] for p in all_programs)
    for cat, count in cat_dist.most_common():
        logger.info(f"  {cat}: {count}")

    logger.info("\n--- Degree Level Distribution ---")
    level_dist = Counter(p["degree_level"] for p in all_programs)
    for lvl, count in level_dist.most_common():
        logger.info(f"  {lvl}: {count}")

    logger.info("\n--- Top 10 Countries ---")
    country_dist = Counter(p["country"] for p in all_programs)
    for c, count in country_dist.most_common(10):
        logger.info(f"  {c}: {count}")

    # Export
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = f"data/exports/training_dataset_{timestamp}.csv"
    json_path = f"data/exports/training_dataset_{timestamp}.json"

    export_csv(all_programs, csv_path)
    export_json(all_programs, json_path)

    # Also save as latest
    export_csv(all_programs, "data/exports/training_dataset_latest.csv")
    export_json(all_programs, "data/exports/training_dataset_latest.json")

    # T5 FIX: Compute and save data hash for reproducibility
    try:
        import pandas as pd

        df_check = pd.read_csv("data/exports/training_dataset_latest.csv")
        csv_bytes = df_check.to_csv(index=False).encode("utf-8")
        data_hash = hashlib.sha256(csv_bytes).hexdigest()[:16]
        manifest = {
            "data_hash": data_hash,
            "timestamp": timestamp,
            "total_programs": len(all_programs),
            "total_universities": len(unis),
            "total_countries": len(countries),
            "csv_path": csv_path,
            "json_path": json_path,
        }
        manifest_path = Path("data/exports/data_manifest.json")
        with open(manifest_path, "w") as f:
            json.dump(manifest, f, indent=2)
        logger.info(f"Data manifest saved: hash={data_hash}")
    except Exception as e:
        logger.warning(f"Could not compute data hash: {e}")

    # Try MongoDB save
    try:
        from src.database.mongodb import get_mongo_connection

        conn = get_mongo_connection()
        db = conn.database

        collection = db["programs"]
        # Bulk insert (remove internal fields)
        clean_programs = []
        for p in all_programs:
            c = {k: v for k, v in p.items() if not k.startswith("_")}
            c["last_updated"] = datetime.now(timezone.utc)
            clean_programs.append(c)

        # Use bulk upsert
        from pymongo import UpdateOne

        ops = []
        for p in clean_programs:
            ops.append(
                UpdateOne(
                    {
                        "university_name": p["university_name"],
                        "program_name": p["program_name"],
                        "degree_type": p["degree_type"],
                    },
                    {"$set": p},
                    upsert=True,
                )
            )

        if ops:
            batch_size = 500
            total_upserted = 0
            total_modified = 0
            for i in range(0, len(ops), batch_size):
                batch = ops[i : i + batch_size]
                result = collection.bulk_write(batch)
                total_upserted += result.upserted_count
                total_modified += result.modified_count
                logger.info(
                    f"  Batch {i//batch_size + 1}: +{result.upserted_count} new, "
                    f"~{result.modified_count} updated"
                )

            logger.info(f"MongoDB: {total_upserted} inserted, {total_modified} updated")

    except Exception as e:
        logger.warning(f"MongoDB save skipped: {e}")
        logger.info("Data exported to CSV/JSON successfully — MongoDB optional")

    logger.info("=" * 60)
    logger.info(f"✅ PHASE 1 COMPLETE: {len(all_programs)} program records generated")
    logger.info(f"   CSV: {csv_path}")
    logger.info(f"   JSON: {json_path}")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
