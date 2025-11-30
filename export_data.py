"""Data Export Script - Export all university and program data from database to various formats."""

import os
import sys
import json
import csv
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

# Setup path FIRST - before any src imports
script_dir = os.path.dirname(os.path.abspath(__file__))
src_dir = os.path.join(script_dir, 'src')
sys.path.insert(0, src_dir)

# Load environment variables
from dotenv import load_dotenv
load_dotenv(os.path.join(script_dir, '.env'))

# Change working directory to src for relative file paths
os.chdir(src_dir)

# Now imports work normally (path is already set up)
from rich.console import Console
from rich.table import Table
from rich.prompt import Prompt, Confirm
from rich.panel import Panel
from rich.text import Text

from database.mongodb import get_mongo_connection
from core.config import get_settings

# Optional imports for Excel export
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

console = Console()


class DataExporter:
    """Export university and program data from MongoDB to various formats."""

    def __init__(self):
        self.settings = get_settings()
        self.mongo_conn = None
        self.programs_data = []
        self.universities_data = []

    def connect_database(self) -> bool:
        """Connect to MongoDB and load data."""
        try:
            self.mongo_conn = get_mongo_connection()

            # Fetch all programs
            programs_collection = self.mongo_conn.collection
            self.programs_data = list(programs_collection.find({}, {'_id': 0}))

            # Fetch all universities
            universities_collection = self.mongo_conn.universities_collection
            self.universities_data = list(universities_collection.find({}, {'_id': 0}))

            console.print(f"[green]✓[/green] Connected to database successfully")
            console.print(f"[blue]📊[/blue] Found {len(self.programs_data)} programs")
            console.print(f"[blue]🏫[/blue] Found {len(self.universities_data)} universities")

            return True

        except Exception as e:
            console.print(f"[red]✗[/red] Failed to connect to database: {e}")
            return False

    def show_data_summary(self):
        """Display summary of loaded data."""
        if not self.programs_data and not self.universities_data:
            console.print("[yellow]⚠[/yellow] No data loaded")
            return

        # Programs summary
        if self.programs_data:
            console.print("\n[bold blue]📊 Programs Data Summary:[/bold blue]")
            universities = set(p.get('university_name', 'Unknown') for p in self.programs_data)
            degree_types = set(p.get('degree_type', 'Unknown') for p in self.programs_data)
            countries = set(p.get('country', 'Unknown') for p in self.programs_data)

            console.print(f"  • Universities: {len(universities)}")
            console.print(f"  • Degree Types: {', '.join(sorted(degree_types))}")
            console.print(f"  • Countries: {', '.join(sorted(countries))}")

        # Universities summary
        if self.universities_data:
            console.print("\n[bold blue]🏫 Universities Data Summary:[/bold blue]")
            countries = set(u.get('country', 'Unknown') for u in self.universities_data)
            tiers = set(u.get('tier', 'Unknown') for u in self.universities_data if u.get('tier'))

            console.print(f"  • Total Universities: {len(self.universities_data)}")
            console.print(f"  • Countries: {', '.join(sorted(countries))}")
            if tiers:
                console.print(f"  • Tiers: {', '.join(sorted(tiers))}")

    def get_export_formats(self) -> List[str]:
        """Get available export formats based on installed packages."""
        formats = ['csv', 'json']

        if OPENPYXL_AVAILABLE or PANDAS_AVAILABLE:
            formats.append('xlsx')

        return formats

    def choose_export_format(self) -> str:
        """Ask user to choose export format."""
        formats = self.get_export_formats()

        console.print("\n[bold cyan]📤 Choose Export Format:[/bold cyan]")

        for i, fmt in enumerate(formats, 1):
            if fmt == 'csv':
                console.print(f"  [{i}] CSV - Comma-separated values (compatible with Excel)")
            elif fmt == 'json':
                console.print(f"  [{i}] JSON - JavaScript Object Notation (structured data)")
            elif fmt == 'xlsx':
                console.print(f"  [{i}] Excel - Microsoft Excel format (.xlsx)")

        while True:
            try:
                choice = Prompt.ask("Enter your choice", choices=[str(i) for i in range(1, len(formats) + 1)])
                return formats[int(choice) - 1]
            except (ValueError, IndexError):
                console.print("[red]Invalid choice. Please try again.[/red]")

    def choose_data_type(self) -> str:
        """Ask user which data to export."""
        console.print("\n[bold cyan]📋 Choose Data to Export:[/bold cyan]")
        console.print("  [1] Programs only")
        console.print("  [2] Universities only")
        console.print("  [3] Both programs and universities")

        while True:
            choice = Prompt.ask("Enter your choice", choices=['1', '2', '3'])
            if choice == '1':
                return 'programs'
            elif choice == '2':
                return 'universities'
            elif choice == '3':
                return 'both'

    def get_programs_field_order(self) -> List[str]:
        """Get prioritized field order for programs data."""
        priority_fields = [
            # Core identification
            'university_name',
            'program_name',
            'degree_type',
            'degree_level',
            'program_category',

            # Basic program info
            'duration_years',
            'tuition_fees',
            'country',
            'city',
            'program_url',

            # Program details
            'description',
            'specializations',
            'prerequisites',

            # Admission requirements
            'application_deadline',
            'gpa_requirement_min',
            'toefl_min',
            'ielts_min',
            'language_requirements',

            # Additional info
            'faculty_research_interests',
            'career_outcomes',
            'admission_requirements',

            # Metadata
            'confidence_score',
            'last_updated',
            'extracted_at',
            'source_url'
        ]
        return priority_fields

    def get_universities_field_order(self) -> List[str]:
        """Get prioritized field order for universities data."""
        priority_fields = [
            # Core identification
            'name',
            'country',
            'city',

            # Rankings and tier
            'tier',
            'qs_world_ranking',
            'the_world_ranking',
            'us_news_ranking',

            # Basic info
            'type',
            'website',
            'description',

            # Metadata
            'updated_at',
            'university_id'
        ]
        return priority_fields

    def get_ordered_fieldnames(self, data: List[Dict], data_type: str) -> List[str]:
        """Get field names in priority order, with remaining fields alphabetically."""
        if data_type == 'programs':
            priority_fields = self.get_programs_field_order()
        elif data_type == 'universities':
            priority_fields = self.get_universities_field_order()
        else:
            # Fallback to alphabetical for unknown types
            all_keys = set()
            for item in data:
                all_keys.update(item.keys())
            return sorted(all_keys)

        # Get all available fields
        all_keys = set()
        for item in data:
            all_keys.update(item.keys())

        # Start with priority fields that exist in the data
        ordered_fields = [field for field in priority_fields if field in all_keys]

        # Add remaining fields alphabetically
        remaining_fields = sorted(all_keys - set(ordered_fields))
        ordered_fields.extend(remaining_fields)

        return ordered_fields

    def get_export_filename(self, data_type: str, format_type: str) -> str:
        """Generate export filename."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{data_type}_{timestamp}.{format_type}"

    def export_csv(self, data: List[Dict], filename: str, data_type: str = 'unknown') -> bool:
        """Export data to CSV format."""
        try:
            if not data:
                console.print("[yellow]⚠[/yellow] No data to export")
                return False

            # Create exports directory if it doesn't exist
            export_dir = Path(self.settings.exports_dir)
            export_dir.mkdir(parents=True, exist_ok=True)

            filepath = export_dir / filename

            # Get ordered field names
            fieldnames = self.get_ordered_fieldnames(data, data_type)

            with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()

                for item in data:
                    # Ensure all fields are present
                    row = {field: item.get(field, '') for field in fieldnames}
                    # Convert complex objects to strings
                    for key, value in row.items():
                        if isinstance(value, (dict, list)):
                            row[key] = json.dumps(value, ensure_ascii=False)
                        elif isinstance(value, datetime):
                            row[key] = value.isoformat()
                    writer.writerow(row)

            console.print(f"[green]✓[/green] Exported {len(data)} records to {filepath}")
            return True

        except Exception as e:
            console.print(f"[red]✗[/red] Failed to export CSV: {e}")
            return False

    def export_json(self, data: List[Dict], filename: str) -> bool:
        """Export data to JSON format."""
        try:
            if not data:
                console.print("[yellow]⚠[/yellow] No data to export")
                return False

            # Create exports directory if it doesn't exist
            export_dir = Path(self.settings.exports_dir)
            export_dir.mkdir(parents=True, exist_ok=True)

            filepath = export_dir / filename

            # Convert datetime objects and complex objects
            serializable_data = []
            for item in data:
                item_copy = item.copy()
                for key, value in item_copy.items():
                    if isinstance(value, datetime):
                        item_copy[key] = value.isoformat()
                    elif isinstance(value, (dict, list)):
                        # Keep as is for JSON
                        pass
                serializable_data.append(item_copy)

            with open(filepath, 'w', encoding='utf-8') as jsonfile:
                json.dump(serializable_data, jsonfile, indent=2, ensure_ascii=False)

            console.print(f"[green]✓[/green] Exported {len(data)} records to {filepath}")
            return True

        except Exception as e:
            console.print(f"[red]✗[/red] Failed to export JSON: {e}")
            return False

    def export_excel(self, data: List[Dict], filename: str, data_type: str = 'unknown') -> bool:
        """Export data to Excel format."""
        try:
            if not data:
                console.print("[yellow]⚠[/yellow] No data to export")
                return False

            # Create exports directory if it doesn't exist
            export_dir = Path(self.settings.exports_dir)
            export_dir.mkdir(parents=True, exist_ok=True)

            filepath = export_dir / filename

            # Try pandas first, fallback to openpyxl
            if PANDAS_AVAILABLE:
                return self._export_excel_pandas(data, filepath, data_type)
            elif OPENPYXL_AVAILABLE:
                return self._export_excel_openpyxl(data, filepath, data_type)
            else:
                console.print("[red]✗[/red] Neither pandas nor openpyxl available for Excel export")
                return False

        except Exception as e:
            console.print(f"[red]✗[/red] Failed to export Excel: {e}")
            return False

    def _export_excel_pandas(self, data: List[Dict], filepath: Path, data_type: str = 'unknown') -> bool:
        """Export using pandas."""
        try:
            df = pd.DataFrame(data)

            # Get ordered columns
            ordered_columns = self.get_ordered_fieldnames(data, data_type)
            # Only include columns that exist in the dataframe
            existing_columns = [col for col in ordered_columns if col in df.columns]
            df = df[existing_columns]

            # Convert datetime columns
            for col in df.columns:
                if df[col].dtype == 'object':
                    # Try to convert datetime strings
                    df[col] = df[col].apply(lambda x: x.isoformat() if isinstance(x, datetime) else x)

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Data', index=False)

                # Basic formatting
                workbook = writer.book
                worksheet = writer.sheets['Data']

                # Header formatting
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font

            console.print(f"[green]✓[/green] Exported {len(data)} records to {filepath} (pandas)")
            return True

        except Exception as e:
            console.print(f"[red]✗[/red] Pandas Excel export failed: {e}")
            return False

    def _export_excel_openpyxl(self, data: List[Dict], filepath: Path, data_type: str = 'unknown') -> bool:
        """Export using openpyxl."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"

            if not data:
                wb.save(filepath)
                return True

            # Get ordered columns
            columns = self.get_ordered_fieldnames(data, data_type)

            # Write header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))

            for col_idx, header in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border

            # Write data
            for row_idx, item in enumerate(data, 2):
                for col_idx, key in enumerate(columns, 1):
                    value = item.get(key, '')
                    if isinstance(value, datetime):
                        value = value.isoformat()
                    elif isinstance(value, (dict, list)):
                        value = json.dumps(value, ensure_ascii=False)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border

            # Auto-adjust column widths
            for col_idx, _ in enumerate(columns, 1):
                max_length = 0
                for row_idx in range(1, len(data) + 2):
                    cell_value = str(ws.cell(row=row_idx, column=col_idx).value or '')
                    max_length = max(max_length, len(cell_value))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

            wb.save(filepath)
            console.print(f"[green]✓[/green] Exported {len(data)} records to {filepath} (openpyxl)")
            return True

        except Exception as e:
            console.print(f"[red]✗[/red] OpenPyXL Excel export failed: {e}")
            return False

    def export_data(self, data_type: str, format_type: str) -> bool:
        """Export data based on type and format."""
        if data_type == 'programs':
            data = self.programs_data
            filename = self.get_export_filename('programs', format_type)
        elif data_type == 'universities':
            data = self.universities_data
            filename = self.get_export_filename('universities', format_type)
        elif data_type == 'both':
            # Export both separately
            programs_success = self.export_data('programs', format_type)
            universities_success = self.export_data('universities', format_type)
            return programs_success and universities_success
        else:
            console.print(f"[red]✗[/red] Unknown data type: {data_type}")
            return False

        if not data:
            console.print(f"[yellow]⚠[/yellow] No {data_type} data to export")
            return False

        # Export based on format
        if format_type == 'csv':
            return self.export_csv(data, filename, data_type)
        elif format_type == 'json':
            return self.export_json(data, filename)
        elif format_type == 'xlsx':
            return self.export_excel(data, filename, data_type)
        else:
            console.print(f"[red]✗[/red] Unknown format: {format_type}")
            return False

    def show_export_location(self):
        """Show where exported files are saved."""
        export_dir = Path(self.settings.exports_dir)
        if export_dir.exists():
            console.print(f"\n[bold green]📁 Export Location:[/bold green] {export_dir.absolute()}")
        else:
            console.print(f"\n[yellow]⚠[/yellow] Export directory will be created at: {export_dir.absolute()}")


def main():
    """Main function to run the data exporter."""
    console.print(Panel.fit(
        "[bold blue]🏫 University Data Exporter[/bold blue]\n"
        "Export university and program data from database to various formats",
        title="📤 Data Export Tool"
    ))

    exporter = DataExporter()

    # Connect to database
    if not exporter.connect_database():
        console.print("[red]❌ Cannot proceed without database connection[/red]")
        return

    # Show data summary
    exporter.show_data_summary()

    # Show export location
    exporter.show_export_location()

    # Choose data type
    data_type = exporter.choose_data_type()

    # Choose format
    format_type = exporter.choose_export_format()

    # Confirm export
    console.print(f"\n[bold yellow]⚡ Ready to export:[/bold yellow]")
    console.print(f"  • Data: {data_type}")
    console.print(f"  • Format: {format_type.upper()}")

    if not Confirm.ask("Proceed with export?"):
        console.print("[yellow]Export cancelled[/yellow]")
        return

    # Perform export
    console.print(f"\n[bold blue]🔄 Exporting {data_type} data to {format_type.upper()}...[/bold blue]")

    success = exporter.export_data(data_type, format_type)

    if success:
        console.print(f"\n[green]🎉 Export completed successfully![/green]")
    else:
        console.print(f"\n[red]❌ Export failed[/red]")


if __name__ == '__main__':
    main()